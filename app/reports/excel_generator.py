import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from sqlalchemy.orm import Session
from app.database.crud_products import get_products  


def generate_report(db: Session, file_name="products_report.xlsx"):
    """
    Gera um relatório estratégico em Excel com:
    - Coloração condicional: 
        🔴 Vermelho para preços > 100
        🟢 Verde para preços ≤ 100
    - Gráfico de barras para análise visual.
    - Melhor formatação dos dados e layout mais limpo.
    """

    # 🔹 1. Obter dados do MySQL e converter em DataFrame
    products = get_products(db)

    if not products:
        print("⚠️ Nenhum produto encontrado no banco de dados! O relatório não será gerado.")
        return

    df = pd.DataFrame([p.__dict__ for p in products])
    print("📌 DataFrame colunas disponíveis:", df.columns.tolist())

    # 🔹 2. Remover colunas técnicas do SQLAlchemy
    # df.drop(columns=["_sa_instance_state"], inplace=True, errors="ignore")

    # 🔹 3. Garantir que as colunas essenciais existem
    if "price" not in df.columns or "category" not in df.columns:
        print("⚠️ As colunas 'price' ou 'category' não foram encontradas! O relatório não pode ser gerado.")
        return

    df["price"] = pd.to_numeric(df["price"], errors="coerce")

    # 🔹 4. Criar um novo arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"

    # 🔹 5. Estilização dos cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 🔹 6. Adicionar dados ao Excel
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # 🔹 7. Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # 🔹 8. Aplicar coloração condicional:
    # 🔴 Vermelho para preços > 100
    # 🟢 Verde para preços ≤ 100
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):  # Coluna de preços
        for cell in row:
            try:
                if isinstance(cell.value, (int, float)):  # Certifica que é numérico
                    if cell.value > 100:
                        cell.fill = red_fill  # Preço acima de 100 fica vermelho
                    else:
                        cell.fill = green_fill  # Preço até 100 fica verde
            except (ValueError, TypeError):
                pass 

    # 🔹 9. Criar aba extra para estatísticas
    ws_stats = wb.create_sheet(title="Estatísticas")

    if "price" in df.columns and "category" in df.columns:
        stats = df.groupby("category")["price"].agg(["mean", "max", "min"]).reset_index()
        stats.rename(columns={"mean": "Média", "max": "Máximo", "min": "Mínimo"}, inplace=True)

        # Adicionar cabeçalhos formatados
        for col_num, column_title in enumerate(stats.columns, 1):
            cell = ws_stats.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Adicionar dados ao Excel
        for row in dataframe_to_rows(stats, index=False, header=False):
            ws_stats.append(row)

        # 🔹 10. Criar gráfico de barras para análise visual
        chart = BarChart()
        chart.title = "Comparação de Preços por Categoria"
        chart.x_axis.title = "Categoria"
        chart.y_axis.title = "Preço Médio"
        chart.style = 13

        data = Reference(ws_stats, min_col=2, min_row=1, max_row=ws_stats.max_row, max_col=2)
        categories = Reference(ws_stats, min_col=1, min_row=2, max_row=ws_stats.max_row)
        series = Series(data, title_from_data=True)
        series.graphicalProperties.solidFill = "4472C4"  
        chart.series.append(series)

        ws_stats.add_chart(chart, "E5")  

    else:
        print("⚠️ Não foi possível gerar estatísticas porque 'category' ou 'price' não existem no DataFrame.")

    # 🔹 11. Salvar o arquivo Excel
    wb.save(file_name)

    print(f"✅ Relatório estratégico gerado com sucesso: {file_name}")
    return {"message": f"Relatório gerado: {file_name}"}
