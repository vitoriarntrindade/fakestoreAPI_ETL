import pytest
import pandas as pd
import os
from openpyxl import load_workbook
from unittest.mock import MagicMock
import factory

from app.reports.excel_generator import generate_report
from app.database.model_product import Product


class ProductFactory(factory.Factory):
    """Fábrica para criar instâncias de produtos fictícios."""

    class Meta:
        model = Product

    id = factory.Sequence(lambda n: n + 1)
    name = factory.Faker("word")
    category = factory.Iterator(["Eletrônicos", "Móveis", "Roupas", "Alimentos"])
    price = factory.Faker("random_int", min=10, max=200)


@pytest.fixture
def fake_db_session():
    """
    Mock da sessão do banco de dados, simulando uma consulta SQL.
    Retorna uma lista fictícia de produtos para evitar chamadas reais ao banco.
    """
    session = MagicMock()
    session.query().all.return_value = [
        ProductFactory() for _ in range(10)
    ]
    return session


@pytest.fixture
def excel_file(tmp_path, fake_db_session):
    """
    Gera um arquivo Excel de teste baseado nos dados fictícios fornecidos.

    Args:
        tmp_path: Diretório temporário onde o arquivo será salvo.
        fake_db_session: Sessão simulada do banco de dados.

    Returns:
        Path do arquivo Excel gerado.
    """
    file_path = tmp_path / "test_products_report.xlsx"
    generate_report(fake_db_session, str(file_path))
    return file_path


def test_excel_file_created(excel_file):
    """Verifica se o arquivo Excel foi gerado com sucesso."""
    assert os.path.exists(excel_file), "O arquivo Excel não foi gerado!"


def test_excel_has_correct_columns(excel_file):
    """
    Garante que o arquivo Excel contém todas as colunas obrigatórias.

    Valida se as colunas ['id', 'name', 'category', 'price'] estão na planilha "Produtos".
    """
    wb = load_workbook(excel_file)
    ws = wb["Produtos"]

    expected_columns = ["id", "name", "category", "price"]
    actual_columns = [cell.value for cell in ws[1]]  # Primeira linha contém os cabeçalhos

    assert set(expected_columns) == set(
        actual_columns), f"Colunas esperadas: {expected_columns}, encontradas: {actual_columns}"


def test_price_column_formatting(excel_file):
    """
    Valida se a formatação da coluna 'price' foi aplicada corretamente.

    - Preços > 100 → 🔴 Vermelho (`FFFF0000`, `00FF0000`, `FF0000`)
    - Preços ≤ 100 → 🟢 Verde (`FF00FF00`, `0000FF00`, `00FF00`)
    """
    wb = load_workbook(excel_file)
    ws = wb["Produtos"]

    RED_COLORS = {"FFFF0000", "00FF0000", "FF0000"}  # Possíveis formatos de vermelho
    GREEN_COLORS = {"FF00FF00", "0000FF00", "00FF00"}  # Possíveis formatos de verde

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            price = cell.value
            color = cell.fill.start_color.index  # Código de cor da célula

            if price > 100:
                assert color in RED_COLORS, f"Erro: {price} deveria estar em vermelho, mas tem cor {color}"
            else:
                assert color in GREEN_COLORS, f"Erro: {price} deveria estar em verde, mas tem cor {color}"


def test_statistics_sheet_exists(excel_file):
    """Confirma se a aba de estatísticas foi criada corretamente no arquivo Excel."""
    wb = load_workbook(excel_file)
    assert "Estatísticas" in wb.sheetnames, "Erro: A aba 'Estatísticas' não foi encontrada!"


def test_statistics_calculation(excel_file):
    """
    Valida se os cálculos estatísticos foram corretamente aplicados.

    Confirma que os valores na aba 'Estatísticas' são números válidos.
    """
    wb = load_workbook(excel_file)
    ws = wb["Estatísticas"]

    values = [row[0] for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2, values_only=True)]

    assert all(isinstance(v, (int, float)) for v in values), "Erro: A coluna estatística contém valores inválidos!"
    assert len(values) > 0, "Erro: Nenhum valor foi encontrado na coluna estatística!"


def test_chart_exists(excel_file):
    """Verifica se o gráfico de barras foi adicionado corretamente na aba 'Estatísticas'."""
    wb = load_workbook(excel_file)
    ws = wb["Estatísticas"]

    assert len(ws._charts) > 0, "Erro: Nenhum gráfico encontrado na aba 'Estatísticas'!"
